from pathlib import Path
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from get_recently_played import get_tracklist
from get_last_hour import last_hour

hour_ago_epoch = last_hour()
tracks = get_tracklist(hour_ago_epoch)

TODAY = str(datetime.today().date())
starting_hour = time.strftime("%H%M", time.localtime(hour_ago_epoch / 1000))


def make_workbook():
    # if workbook exists, open it
    workbook_path = f'./data/{str(TODAY)}.xlsx'
    worksheet_name = str(starting_hour)
    workbook_location = Path(workbook_path)
    # Check if workbook exists
    if workbook_location.is_file():
        wb = load_workbook(filename=workbook_path)
        # Check if the worksheet already exists
        if worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            ws = wb.create_sheet(worksheet_name)
        return wb, ws
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = worksheet_name
        wb.save(workbook_path)
        return wb, ws


def write_workbook(wb, ws):
    # write info to workbook
    for x in range(1, len(tracks) + 1):
        ws[f"A{x}"] = str(tracks[x - 1]['song'])
        ws[f"B{x}"] = str(tracks[x - 1]['artist'])
        ws[f"C{x}"] = str(tracks[x - 1]['album'])
        ws[f"D{x}"] = str(tracks[x - 1]['duration'])
        ws[f"E{x}"] = str(tracks[x - 1]['release_date'])
        ws[f"F{x}"] = str(tracks[x - 1]['played_at'])
        ws[f"G{x}"] = str(tracks[x - 1]['song_popularity'])
    wb.save(f'./data/{str(TODAY)}.xlsx')


if __name__ == '__main__':
    workbook, worksheet = make_workbook()
    write_workbook(workbook, worksheet)
